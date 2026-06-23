#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, io
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
"""
gerar_metas.py
Lê os arquivos da pasta do mês + config e gera:
  - consolidado.xlsx  (auditoria / prova real)
  - dashboard_metas.html (painel visual)
"""

import os
import re
import secrets
import unicodedata
import hashlib
import subprocess
import pandas as pd
import json
import webbrowser
import warnings

def _abrir_chrome(url: str):
    """Tenta abrir Chrome. Se não encontrar, usa o browser padrão."""
    chrome_paths = [
        r'C:\Program Files\Google\Chrome\Application\chrome.exe',
        r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
        os.path.join(os.environ.get('LOCALAPPDATA', ''),
                     r'Google\Chrome\Application\chrome.exe'),
    ]
    for p in chrome_paths:
        if os.path.exists(p):
            subprocess.Popen([p, url])
            return
    webbrowser.open(url)   # fallback
from pathlib import Path
from datetime import date, datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

SISTEMA_DIR   = Path(__file__).parent          # pasta _sistema/ (scripts e templates)
BASE_DIR      = SISTEMA_DIR.parent             # pasta raiz (Excels de entrada, index/app.html, git)
CONFIG_DIR    = BASE_DIR / "config"
SAIDAS_DIR    = BASE_DIR / "_saidas"           # saidas geradas (dashboard local + consolidado)
SAIDAS_DIR.mkdir(exist_ok=True)
TEMPLATE      = SISTEMA_DIR / "dashboard_template_metas.html"
LOGIN_TEMPLATE = SISTEMA_DIR / "login_template.html"
OUTPUT_HTML   = SAIDAS_DIR / "dashboard_metas.html"
OUTPUT_XLS    = SAIDAS_DIR / "consolidado.xlsx"
USUARIOS_FILE = SISTEMA_DIR / "usuarios.json"
LOGIN_OUTPUT  = BASE_DIR / "index.html"
SUPER_DIR     = BASE_DIR / "super"
DIG_XLSX      = BASE_DIR / "base_digitacoes.xlsx"   # Excel de digitações
HIER_XLSX     = BASE_DIR / "HIERARQUIA COMERCIAL.xlsx"  # Hierarquia comercial externa

# ── Configuração Digitações ──────────────────────────────────────────────────
DIAS_JANELA_DIG = 10
BANCOS_ESTRATEGICOS_DIG = [
    '2S CONSIG', 'AMIGOZ', 'BANCO C6 BANK', 'BANCO DAYCOVAL',
    'BANCO DIGIO S.A.', 'CAPITAL CONSIG SCD S.A.', 'FACTA FINANCEIRA',
]
SUPERS_EXCLUIDOS_DIG = ['DINHO SILVA']

# ── Configuração GitHub ──────────────────────────────────────────────────────
PUSH_GIT = True   # Mude para False para desabilitar o push automático

FERIADOS_2026 = {
    date(2026, 1, 1), date(2026, 4, 3), date(2026, 4, 21),
    date(2026, 5, 1), date(2026, 6, 4), date(2026, 9, 7),
    date(2026, 10, 12), date(2026, 11, 2), date(2026, 11, 20),
    date(2026, 12, 25),
}

MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
            "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
MESES_ABR = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]

# ─────────────────────────────────────────────────────────────────────────────
# Digitações — leitura e preparação dos dados
# ─────────────────────────────────────────────────────────────────────────────

def _dig_encontrar_col(df, *palavras):
    for col in df.columns:
        cl = col.lower()
        if all(p in cl for p in palavras):
            return col
    return None

def _parse_valor_br(val):
    """
    Converte o valor de digitação (relatório Nova Financeira) para float.
    O relatório vem em formato BR: ponto = milhar, vírgula = decimal.
      '7.448,22' -> 7448.22
      '0,00'     -> 0.0
      '150.00'   -> 150.0   (caso já venha em ponto decimal)
      '' / None  -> None
    NÃO usa a regra de "inteiro puro = centavos" (diferente do parser de
    comissões) — confirmado que esta coluna sempre vem com decimais explícitos.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if s in ("", "nan", "None", "NaT"):
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def _carregar_hierarquia_externa():
    if not HIER_XLSX.exists():
        return None
    hier = pd.read_excel(HIER_XLSX)
    hier.columns = [str(c).strip() for c in hier.columns]
    col_map = {}
    for c in hier.columns:
        cl = c.lower()
        if ('cod' in cl and 'parceiro' in cl) or cl == 'corretor': col_map[c] = 'CodParceiro'
        elif 'comercial' in cl: col_map[c] = 'Comercial'
        elif 'regional' in cl or 'reginal' in cl: col_map[c] = 'Regional'
        elif 'superinten' in cl: col_map[c] = 'Superintendente'
    hier = hier.rename(columns=col_map)
    needed = ['CodParceiro', 'Comercial', 'Regional', 'Superintendente']
    if all(c in hier.columns for c in needed):
        print(f'  [DIG] Hierarquia externa: {len(hier)} registros')
        return hier[needed]
    hier2 = pd.read_excel(HIER_XLSX, header=None)
    if len(hier2.columns) >= 7:
        print(f'  [DIG] Hierarquia externa (posicional): {len(hier2)} registros')
        return hier2[[0,4,5,6]].rename(columns={0:'CodParceiro',4:'Comercial',5:'Regional',6:'Superintendente'})
    print('  [DIG] AVISO: HIERARQUIA COMERCIAL.xlsx com colunas nao reconhecidas')
    return None

def ler_dados_digitacoes(caminho):
    """Lê base_digitacoes.xlsx (aba Base) e cruza com HIERARQUIA COMERCIAL.xlsx."""
    xl   = pd.ExcelFile(caminho)
    abas = [a.strip() for a in xl.sheet_names]

    if 'Base' not in abas:
        print('  [DIG] ERRO: aba Base nao encontrada em base_digitacoes.xlsx')
        return None

    base = pd.read_excel(xl, sheet_name='Base')
    base.columns = [str(c).strip() for c in base.columns]

    col_corretor = _dig_encontrar_col(base, 'corretor')
    col_banco    = _dig_encontrar_col(base, 'banco', 'empr')
    col_orgao    = _dig_encontrar_col(base, 'convenio') or _dig_encontrar_col(base, 'conv')
    col_operacao = _dig_encontrar_col(base, 'opera')
    col_status   = _dig_encontrar_col(base, 'status')
    col_data     = _dig_encontrar_col(base, 'data', 'digit', 'banco')
    col_valor    = (_dig_encontrar_col(base, 'valor', 'base')
                    or _dig_encontrar_col(base, 'produ', 'bruta')
                    or _dig_encontrar_col(base, 'digit'))

    df = base[[col_corretor, col_banco, col_orgao, col_operacao,
               col_status, col_data, col_valor]].copy()
    df.columns = ['Corretor','Banco','Orgao','Operacao','Status','Data','Digitacao']

    hier = _carregar_hierarquia_externa()
    if hier is not None:
        df['_cod'] = df['Corretor'].astype(str).str.extract(r'^(\d+)', expand=False).fillna('').astype(str)
        hier = hier.copy()
        hier['_cod'] = hier['CodParceiro'].astype(str)
        df = df.merge(hier[['_cod','Comercial','Regional','Superintendente']],
                      on='_cod', how='left')
        df = df.drop(columns=['_cod'])
    else:
        print('  [DIG] AVISO: HIERARQUIA COMERCIAL.xlsx nao encontrado — sem dados de hierarquia')
        df['Comercial']       = None
        df['Regional']        = None
        df['Superintendente'] = None

    return df[['Corretor','Comercial','Regional','Superintendente',
               'Banco','Orgao','Operacao','Status','Data','Digitacao']]

def processar_digitacoes():
    """Retorna (records_json, estrategicos_json, periodo_str) ou ([], [], '')."""
    if not DIG_XLSX.exists():
        print(f"  [DIG] base_digitacoes.xlsx não encontrado — aba de digitações sem dados.")
        return [], json.dumps(BANCOS_ESTRATEGICOS_DIG, ensure_ascii=False), ''
    try:
        from datetime import timedelta as _td
        print(f"  [DIG] Lendo {DIG_XLSX.name}...")
        df = ler_dados_digitacoes(str(DIG_XLSX))
        if df is None or df.empty:
            return [], json.dumps(BANCOS_ESTRATEGICOS_DIG, ensure_ascii=False), ''
        df['Data']      = pd.to_datetime(df['Data'], errors='coerce')
        df['Digitacao'] = pd.to_numeric(df['Digitacao'], errors='coerce').fillna(0)
        df = df.dropna(subset=['Data'])
        if SUPERS_EXCLUIDOS_DIG:
            pat  = '|'.join(SUPERS_EXCLUIDOS_DIG)
            mask = df['Superintendente'].astype(str).str.contains(pat, case=False, na=False)
            df   = df[~mask]
        max_date = df['Data'].max()
        min_date = max_date - _td(days=DIAS_JANELA_DIG - 1)
        df = df[(df['Data'] >= min_date) & (df['Data'] <= max_date)].copy()
        periodo = f"{min_date.strftime('%d/%m/%Y')} a {max_date.strftime('%d/%m/%Y')}"
        print(f"  [DIG] {len(df)} registros | {periodo}")
        def _strip_prefix(s):
            """Remove prefixo numérico tipo '1003 - ' dos campos hierárquicos."""
            if not s:
                return s
            import re as _re
            return _re.sub(r'^\d+\s*[-–]\s*', '', str(s)).strip()

        records = []
        for row in df.to_dict('records'):
            records.append({
                'banco':           str(row['Banco'])          if pd.notna(row['Banco'])          else '',
                'operacao':        str(row['Operacao'])        if pd.notna(row['Operacao'])        else '',
                'status':          str(row['Status']).lower().strip() if pd.notna(row['Status']) else '',
                'superintendente': _strip_prefix(str(row['Superintendente'])) if pd.notna(row['Superintendente']) else '',
                'regional':        _strip_prefix(str(row['Regional']))        if pd.notna(row['Regional'])        else '',
                'comercial':       _strip_prefix(str(row['Comercial']))       if pd.notna(row['Comercial'])       else '',
                'orgao':           str(row['Orgao'])            if pd.notna(row['Orgao'])            else '',
                'data':            row['Data'].strftime('%Y-%m-%d'),
                'valor':           round(float(row['Digitacao']), 2),
            })
        return records, json.dumps(BANCOS_ESTRATEGICOS_DIG, ensure_ascii=False), periodo
    except Exception as e:
        print(f"  [DIG] Erro ao processar digitações: {e}")
        return [], json.dumps(BANCOS_ESTRATEGICOS_DIG, ensure_ascii=False), ''

# ─────────────────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────────────────

def dias_uteis_entre(start: date, end: date, feriados: set) -> int:
    n, cur = 0, start
    while cur <= end:
        if cur.weekday() < 5 and cur not in feriados:
            n += 1
        cur += timedelta(days=1)
    return n

def clean_name(s: str) -> str:
    s = str(s).strip()
    return s.split(" - ", 1)[1].strip() if " - " in s else s

def base_name(s: str) -> str:
    """Remove sufixos entre parênteses e retorna nome em maiúsculas para comparação."""
    return re.sub(r'\s*\([^)]+\)\s*', '', str(s)).strip().upper()

def normalize(s: str) -> str:
    """Remove acentos e retorna em maiúsculas — para comparações sem sensibilidade a acentuação."""
    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii').upper().strip()

def parenthetical(s: str) -> str:
    """Retorna o conteúdo dentro dos parênteses (normalizado), ou vazio."""
    m = re.search(r'\(([^)]+)\)', str(s))
    return normalize(m.group(1)) if m else ""

def eh_dinho(s: str) -> bool:
    return "DINHO" in str(s).upper()

def fmt(v, casas=2):
    return round(float(v), casas) if v is not None and str(v) not in ("nan","None","") else None

def pct(num, den):
    return round(num / den * 100, 1) if den and den > 0 else None

def gap(proj, meta):
    return round(proj - meta, 2) if meta and meta > 0 else None

# ─────────────────────────────────────────────────────────────────────────────
# Carregamento
# ─────────────────────────────────────────────────────────────────────────────

def load_excel_or_csv(folder: Path, stem: str) -> pd.DataFrame:
    for ext in [".xlsx", ".xls", ".csv"]:
        p = folder / f"{stem}{ext}"
        if p.exists():
            df = pd.read_csv(p) if ext == ".csv" else pd.read_excel(p, header=0)
            df.columns = [str(c).strip() for c in df.columns]
            return df
    raise FileNotFoundError(f"Arquivo '{stem}' não encontrado em {folder}")

def _parse_mes_ano(v) -> str:
    """Normaliza Mês/Ano para 'YYYY-MM': aceita '05/2026', '2026-05', Timestamp."""
    try:
        ts = pd.Timestamp(v)
        if pd.notna(ts) and ts.year > 2000:
            return f"{ts.year:04d}-{ts.month:02d}"
    except Exception:
        pass
    s = str(v).strip()
    for sep in ["/", "-"]:
        parts = s.split(sep)
        if len(parts) == 2:
            a, b = parts[0].strip(), parts[1].strip()
            if len(b) == 4 and b.isdigit() and a.isdigit():
                return f"{b}-{int(a):02d}"   # MM/YYYY
            if len(a) == 4 and a.isdigit() and len(b) == 2 and b.isdigit():
                return f"{a}-{b}"            # YYYY-MM
    return ""

def _load_producao_anual() -> pd.DataFrame:
    for fname in ["producao_2026", "producao"]:
        for ext in [".xlsx", ".xls", ".csv"]:
            p = BASE_DIR / f"{fname}{ext}"
            if p.exists():
                df = pd.read_csv(p) if ext == ".csv" else pd.read_excel(p, header=0)
                df.columns = [str(c).strip() for c in df.columns]
                return df
    raise FileNotFoundError(
        "Arquivo de producao nao encontrado.\n"
        "Esperado: producao_2026.xlsx na pasta do projeto."
    )

def _load_meta_banco_v2():
    """Carrega meta_banco.xlsx com abas 'metas' e 'comerciais'.
    Retorna (metas_df, coms_df).
    """
    for ext in [".xlsx", ".xls"]:
        p = BASE_DIR / f"meta_banco{ext}"
        if p.exists():
            sheets = pd.read_excel(p, sheet_name=None, header=1)  # linha 2 é o header (linha 1 é instrução)
            def norm_cols(df):
                df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
                return df
            metas_df = norm_cols(sheets.get("metas", pd.DataFrame()))
            coms_df  = norm_cols(sheets.get("comerciais", pd.DataFrame(columns=["mes","comercial"])))
            return metas_df, coms_df
    # fallback: sem arquivo
    return pd.DataFrame(), pd.DataFrame(columns=["mes","comercial"])

# ─────────────────────────────────────────────────────────────────────────────
# Processamento principal
# ─────────────────────────────────────────────────────────────────────────────

def processar():
    today  = date.today()
    ontem  = today - timedelta(days=1)

    # ── Config ────────────────────────────────────────────────────────────────
    cal_df = load_excel_or_csv(CONFIG_DIR, "calendario")
    mg_df  = load_excel_or_csv(CONFIG_DIR, "meta_global_2026")

    # ── Mês atual (detectado por hoje) ────────────────────────────────────────
    ano, mes = today.year, today.month
    mes_str  = f"{ano:04d}-{mes:02d}"

    # ── Dias úteis ────────────────────────────────────────────────────────────
    cal_row = cal_df[cal_df["mes"].astype(str) == mes_str]
    if cal_row.empty:
        raise ValueError(f"Mes {mes_str} nao encontrado em config/calendario.xlsx")
    du_total    = int(cal_row["dias_uteis"].iloc[0])
    feriados    = FERIADOS_2026 if ano == 2026 else set()
    primeiro    = date(ano, mes, 1)
    du_passados = max(1, dias_uteis_entre(primeiro, ontem, feriados))
    fator       = du_total / du_passados

    print(f"  Mes: {mes_str}  |  DU total: {du_total}  |  DU passados: {du_passados}  |  Fator: {fator:.4f}")

    # ── Producao (arquivo unico anual) ─────────────────────────────────────────
    prod_all = _load_producao_anual()
    col_mes  = next((c for c in prod_all.columns
                     if "ês" in c or "es/" in c.lower() or c.lower() == "mes"), None)
    if col_mes is None:
        raise ValueError("Coluna 'Mes/Ano' nao encontrada no arquivo de producao")
    prod_all["_mes_str"] = prod_all[col_mes].apply(_parse_mes_ano)
    prod = prod_all[prod_all["_mes_str"] == mes_str].copy()
    if prod.empty:
        raise ValueError(
            f"Nenhuma producao para {mes_str} em producao_2026.xlsx.\n"
            f"Verifique coluna Mes/Ano (formato: MM/AAAA ou AAAA-MM)."
        )

    # Detecta colunas-chave
    col_prd = next((c for c in prod.columns if "Valor" in c or "Soma" in c), None)
    col_ban = next((c for c in prod.columns if c == "Banco"), None)
    col_sup = next((c for c in prod.columns if "Super" in c or "super" in c), None)
    col_reg = next((c for c in prod.columns if "Regional" in c), None)
    col_com = next((c for c in prod.columns if "Comercial" in c), None)

    for col, nome in [(col_prd,"Soma Valor Base"),(col_ban,"Banco"),
                      (col_sup,"Superintendente"),(col_reg,"Regional"),(col_com,"Comercial")]:
        if col is None:
            raise ValueError(f"Coluna '{nome}' não encontrada em producao.xlsx")

    prod["_eh_dinho"]  = prod[col_sup].apply(eh_dinho)
    prod["_super"]     = prod[col_sup].apply(clean_name)
    prod["_regional"]  = prod[col_reg].apply(clean_name)
    prod["_comercial"] = prod[col_com].apply(clean_name)
    prod["_prod"]      = pd.to_numeric(prod[col_prd], errors="coerce").fillna(0)
    prod["_banco_raw"] = prod[col_ban].astype(str).str.upper()

    # Sets normalizados para comparação (sem acentos)
    super_base_n = {normalize(base_name(n)) for n in prod["_super"].dropna().unique()}
    reg_base_n   = {normalize(base_name(n)) for n in prod["_regional"].dropna().unique()}

    def _com_tem_meta(com_name: str) -> bool:
        # Nome base bate com super ou regional → sem meta
        if normalize(base_name(com_name)) in super_base_n | reg_base_n:
            return False
        # Sufixo entre parênteses bate com nome base de alguma regional
        # ex: "PEDRO (MELHOR CRÉDITO)" → "(MELHOR CRÉDITO)" bate com regional "MELHOR CRÉDITO (REGIONAL)"
        par = parenthetical(com_name)
        if par and par in reg_base_n:
            return False
        return True

    prod["_tem_meta"] = prod["_comercial"].apply(_com_tem_meta)

    # Dinho: meta só no nível super — regionais e comerciais abaixo dele sem meta individual
    prod.loc[prod["_eh_dinho"], "_tem_meta"] = False

    # ── Colunas auxiliares para filtros condicionais de banco ─────────────────
    col_con = next((c for c in prod.columns if "onv" in c.lower()), None)
    col_tip = next((c for c in prod.columns if "ipo" in c.lower() and "op" in c.lower()), None)
    col_par = next((c for c in prod.columns if "arceiro" in c.lower()), None)
    col_sta = next((c for c in prod.columns if "tatus" in c.lower()), None)
    prod["_convenio_raw"] = prod[col_con].astype(str).str.upper() if col_con else ""
    prod["_tipo_raw"]     = prod[col_tip].astype(str).str.upper() if col_tip else ""

    # ── Meta banco — carrega config e determina mês ativo ────────────────────
    mb_all, coms_all = _load_meta_banco_v2()

    # Mês ativo do banco = último mês presente na aba 'metas'
    # → freeze automático: só muda quando o usuário adicionar novo mês no arquivo
    mes_list_banco = sorted(mb_all["mes"].astype(str).str.strip().unique()) if not mb_all.empty else []
    banco_mes      = mes_list_banco[-1] if mes_list_banco else mes_str
    banco_ano_n    = int(banco_mes[:4])
    banco_mes_n    = int(banco_mes[5:7])

    # Produção do mês do banco (pode diferir do mês global se congelado)
    prod_b = prod_all[prod_all["_mes_str"] == banco_mes].copy()
    if not prod_b.empty:
        prod_b["_super"]         = prod_b[col_sup].apply(clean_name)
        prod_b["_regional"]      = prod_b[col_reg].apply(clean_name)
        prod_b["_comercial"]     = prod_b[col_com].apply(clean_name)
        prod_b["_prod"]          = pd.to_numeric(prod_b[col_prd], errors="coerce").fillna(0)
        prod_b["_banco_raw"]     = prod_b[col_ban].astype(str).str.upper()
        prod_b["_eh_dinho"]      = prod_b[col_sup].apply(eh_dinho)
        prod_b["_convenio_raw"]  = prod_b[col_con].astype(str).str.upper() if col_con else ""
        prod_b["_tipo_raw"]      = prod_b[col_tip].astype(str).str.upper() if col_tip else ""
    else:
        prod_b = prod.copy()  # fallback

    # Fator do mês do banco (mês passado → fator=1.0; mês atual → fator normal)
    import calendar as _cal
    banco_ultimo = date(banco_ano_n, banco_mes_n, _cal.monthrange(banco_ano_n, banco_mes_n)[1])
    if banco_mes == mes_str:
        fator_banco = fator           # mês em curso: usa fator real
    elif today > banco_ultimo:
        fator_banco = 1.0             # mês encerrado: produção = realizado
    else:
        fator_banco = fator           # mês futuro pré-configurado: usa fator corrente

    # Bancos: filtra pelo mês ativo
    mb_df  = mb_all[mb_all["mes"].astype(str).str.strip() == banco_mes].copy() if not mb_all.empty else pd.DataFrame()

    def _parse_filter(val, fallback: str = "") -> str:
        """Converte célula do Excel em padrão regex para str.contains().
        NaN/vazio  → fallback (padrão: '' = sem filtro).
        Múltiplos valores separados por ';'  → padrão OR (ex: 'INSS|FGTS').
        """
        if pd.isna(val):
            return fallback
        s = str(val).strip().upper()
        if not s:
            return fallback
        parts = [re.escape(p.strip()) for p in s.split(";") if p.strip()]
        return "|".join(parts) if parts else fallback

    bancos = []
    for r in mb_df.to_dict('records'):
        nome   = str(r.get("banco_display", "")).strip()
        filtro = _parse_filter(r.get("banco_filtro"), fallback=re.escape(nome.upper()))
        conv_f = _parse_filter(r.get("convenio_filtro"), fallback="")
        tipo_f = _parse_filter(r.get("tipo_filtro"),    fallback="")
        meta   = float(r.get("meta_total", 0) or 0)
        if nome and meta > 0:
            bancos.append({"nome": nome, "filtro": filtro,
                           "conv_f": conv_f, "tipo_f": tipo_f, "meta": meta})
    total_meta_banco = sum(b["meta"] for b in bancos)

    # Comerciais com meta banco — lista explícita da aba 'comerciais'
    coms_mes = coms_all[coms_all["mes"].astype(str).str.strip() == banco_mes] if not coms_all.empty else pd.DataFrame(columns=["mes","comercial"])
    if coms_mes.empty and not coms_all.empty:
        _ult_mes_coms = sorted(coms_all["mes"].astype(str).str.strip().unique())[-1]
        coms_mes = coms_all[coms_all["mes"].astype(str).str.strip() == _ult_mes_coms]
        print(f"  [AVISO] aba comerciais do meta_banco.xlsx sem linhas para {banco_mes} — usando lista de {_ult_mes_coms}. Atualize a planilha!")
    banco_coms_norm = {normalize(clean_name(str(c).strip())) for c in coms_mes.get("comercial", pd.Series()).dropna().unique()}

    # ── Meta global por super (mês) ───────────────────────────────────────────
    mes_abr = MESES_ABR[mes - 1]
    meta_super = {}   # nome_super → meta_mensal
    _mg_first_col = mg_df.columns[0] if not mg_df.empty else None
    for r in mg_df.to_dict('records'):
        nome = clean_name(str(r[_mg_first_col]).strip())
        val  = pd.to_numeric(r.get(mes_abr), errors="coerce")
        if pd.notna(val) and val > 0:
            meta_super[nome] = float(val)

    # Normaliza chave do Dinho: meta_global usa "Sala Dinho", prod usa "DINHO SILVA"
    dinho_prod_names = prod[prod["_eh_dinho"]]["_super"].unique()
    dinho_meta_key   = next((k for k in list(meta_super) if "DINHO" in k.upper() or "SALA" in k.upper()), None)
    if dinho_meta_key and len(dinho_prod_names) > 0:
        dinho_prod_name = str(dinho_prod_names[0])
        if dinho_prod_name != dinho_meta_key:
            meta_super[dinho_prod_name] = meta_super.pop(dinho_meta_key)

    # ── Contagem de comerciais com meta global ────────────────────────────────
    com_meta_df = prod[prod["_tem_meta"] & ~prod["_eh_dinho"]]
    uniq_super_com = (com_meta_df
                      .groupby(["_super", "_comercial"])
                      .size().reset_index()[["_super", "_comercial"]])
    n_com_por_super = uniq_super_com.groupby("_super").size().to_dict()
    n_com_total     = len(uniq_super_com["_comercial"].unique())

    # ── Contagem de comerciais com meta banco (lista explícita) ──────────────
    banco_com_mask = prod_b["_comercial"].apply(lambda c: normalize(c) in banco_coms_norm) & ~prod_b["_eh_dinho"]
    uniq_banco_com = (prod_b[banco_com_mask]
                      .groupby(["_super", "_comercial"])
                      .size().reset_index()[["_super", "_comercial"]])
    n_com_banco_por_super = uniq_banco_com.groupby("_super").size().to_dict()
    n_com_banco_total     = len(banco_coms_norm) or 1  # total da lista config (não da produção)

    print(f"  Comerciais meta global: {n_com_total}  |  meta banco: {n_com_banco_total}  "
          f"(mês banco: {banco_mes})  |  Meta banco/com: R${total_meta_banco/n_com_banco_total:,.0f}")

    # Meta banco por comercial (divisão igual entre os listados)
    meta_bco_por_com = {b["nome"]: b["meta"] / n_com_banco_total for b in bancos}

    # Meta global por comercial = meta_super / n_com_do_super
    def meta_global_com(super_nome: str) -> float:
        n = n_com_por_super.get(super_nome, 1)
        m = meta_super.get(super_nome, 0)
        return m / n if n > 0 and m > 0 else 0

    # ── Produção por banco com filtros condicionais ───────────────────────────
    def prod_banco_cond(b_config: dict, subset: pd.DataFrame) -> float:
        """Filtra subset pelo banco + opcionalmente convênio e tipo operação."""
        mask = subset["_banco_raw"].str.contains(b_config["filtro"], na=False)
        if b_config.get("conv_f"):
            mask &= subset["_convenio_raw"].str.contains(b_config["conv_f"], na=False)
        if b_config.get("tipo_f"):
            mask &= subset["_tipo_raw"].str.contains(b_config["tipo_f"], na=False)
        return float(subset.loc[mask, "_prod"].sum())

    # ── Montagem da hierarquia ────────────────────────────────────────────────
    rows_com  = []   # para consolidado: comerciais × banco
    rows_sum  = []   # para consolidado: comerciais resumo

    supers_out = []

    for sup_raw in sorted(prod[col_sup].dropna().unique(), key=str):
        sup_nome  = clean_name(str(sup_raw))
        sup_df    = prod[prod["_super"] == sup_nome]          # mês global
        sup_df_b  = prod_b[prod_b["_super"] == sup_nome]      # mês banco
        dinho_s   = eh_dinho(sup_raw)

        prod_s   = float(sup_df["_prod"].sum())
        proj_s   = prod_s * fator
        meta_s   = meta_super.get(sup_nome, 0)

        # bancos do super — usa mês banco e n_com_banco
        n_com_banco_s = n_com_banco_por_super.get(sup_nome, 0)
        bancos_sup   = []
        for b in bancos:
            p   = prod_banco_cond(b, sup_df_b)
            m   = 0 if dinho_s else meta_bco_por_com.get(b["nome"], 0) * n_com_banco_s
            pj  = p * fator_banco
            bancos_sup.append({"nome": b["nome"], "meta": fmt(m), "prod": fmt(p),
                                "proj": fmt(pj), "pct": pct(pj, m), "gap": gap(pj, m)})
        meta_banco_s = sum(x["meta"] or 0 for x in bancos_sup)
        prod_banco_s = sum(x["prod"] or 0 for x in bancos_sup)
        proj_banco_s = prod_banco_s * fator_banco

        regionais_out = []

        for reg_raw in sorted(sup_df[col_reg].dropna().unique(), key=str):
            reg_nome  = clean_name(str(reg_raw))
            reg_df    = sup_df[sup_df["_regional"] == reg_nome]       # mês global
            reg_df_b  = sup_df_b[sup_df_b["_regional"] == reg_nome]  # mês banco
            est_reg   = not bool(reg_df["_tem_meta"].any())

            prod_r   = float(reg_df["_prod"].sum())
            proj_r   = prod_r * fator
            n_com_r  = len(reg_df[reg_df["_tem_meta"]]["_comercial"].unique())
            meta_r   = 0 if dinho_s else meta_global_com(sup_nome) * n_com_r

            # n_com_banco para esta regional (comerciais da lista explícita)
            n_com_banco_r = len(
                reg_df_b[reg_df_b["_comercial"].apply(lambda c: normalize(c) in banco_coms_norm)]
                ["_comercial"].unique()
            )
            est_reg_banco = (n_com_banco_r == 0)  # regional sem comerciais na lista banco

            bancos_reg = []
            for b in bancos:
                p  = prod_banco_cond(b, reg_df_b)
                m  = 0 if (dinho_s or est_reg_banco) else meta_bco_por_com.get(b["nome"], 0) * n_com_banco_r
                pj = p * fator_banco
                bancos_reg.append({"nome": b["nome"], "meta": fmt(m), "prod": fmt(p),
                                   "proj": fmt(pj), "pct": pct(pj, m), "gap": gap(pj, m)})
            meta_banco_r = sum(x["meta"] or 0 for x in bancos_reg)
            prod_banco_r = sum(x["prod"] or 0 for x in bancos_reg)
            proj_banco_r = prod_banco_r * fator_banco

            comerciais_out = []

            for com_raw in sorted(reg_df[col_com].dropna().unique(), key=str):
                com_nome  = clean_name(str(com_raw))
                com_df    = reg_df[reg_df["_comercial"] == com_nome]
                com_df_b  = reg_df_b[reg_df_b["_comercial"] == com_nome]
                tem_meta  = bool(com_df["_tem_meta"].any())
                # Meta banco: usa lista explícita (independente do _tem_meta global)
                tem_meta_banco = normalize(com_nome) in banco_coms_norm

                prod_c    = float(com_df["_prod"].sum())
                proj_c    = prod_c * fator
                meta_c    = meta_global_com(sup_nome) if tem_meta else 0

                bancos_com = []
                for b in bancos:
                    p  = prod_banco_cond(b, com_df_b)
                    m  = meta_bco_por_com.get(b["nome"], 0) if tem_meta_banco else 0
                    pj = p * fator_banco
                    bancos_com.append({"nome": b["nome"], "meta": fmt(m), "prod": fmt(p),
                                       "proj": fmt(pj), "pct": pct(pj, m), "gap": gap(pj, m)})
                meta_banco_c = sum(x["meta"] or 0 for x in bancos_com)
                prod_banco_c = sum(x["prod"] or 0 for x in bancos_com)
                proj_banco_c = prod_banco_c * fator_banco

                comerciais_out.append({
                    "nome": com_nome, "tem_meta": tem_meta, "tem_meta_banco": tem_meta_banco,
                    "meta_global": fmt(meta_c), "prod_total": fmt(prod_c),
                    "proj_total": fmt(proj_c),
                    "pct_global": pct(proj_c, meta_c), "gap_global": gap(proj_c, meta_c),
                    "meta_banco_total": fmt(meta_banco_c),
                    "prod_banco_total": fmt(prod_banco_c),
                    "proj_banco_total": fmt(proj_banco_c),
                    "pct_banco": pct(proj_banco_c, meta_banco_c),
                    "bancos": bancos_com,
                })

                # linha consolidado granular
                for bc in bancos_com:
                    rows_com.append({
                        "Superintendente": sup_nome, "Regional": reg_nome,
                        "Comercial": com_nome, "Tem Meta": "Sim" if tem_meta else "Não",
                        "Banco": bc["nome"], "Meta": bc["meta"], "Producao": bc["prod"],
                        "Projecao": bc["proj"], "Pct Ating (%)": bc["pct"], "GAP": bc["gap"],
                    })
                # linha consolidado resumo comercial
                rows_sum.append({
                    "Superintendente": sup_nome, "Regional": reg_nome,
                    "Comercial": com_nome, "Tem Meta": "Sim" if tem_meta else "Não",
                    "Meta Global": fmt(meta_c), "Producao Total": fmt(prod_c),
                    "Projecao Total": fmt(proj_c),
                    "Pct Global (%)": pct(proj_c, meta_c), "GAP Global": gap(proj_c, meta_c),
                    "Meta Banco Total": fmt(meta_banco_c),
                    "Prod Banco Total": fmt(prod_banco_c),
                    "Proj Banco Total": fmt(proj_banco_c),
                    "Pct Banco (%)": pct(proj_banco_c, meta_banco_c),
                })

            regionais_out.append({
                "nome": reg_nome, "eh_estrategico": est_reg,
                "meta_global": fmt(meta_r), "prod_total": fmt(prod_r),
                "proj_total": fmt(proj_r),
                "pct_global": pct(proj_r, meta_r), "gap_global": gap(proj_r, meta_r),
                "meta_banco_total": fmt(meta_banco_r),
                "prod_banco_total": fmt(prod_banco_r),
                "proj_banco_total": fmt(proj_banco_r),
                "pct_banco": pct(proj_banco_r, meta_banco_r),
                "bancos": bancos_reg, "comerciais": comerciais_out,
            })

        supers_out.append({
            "nome": sup_nome, "eh_dinho": dinho_s,
            "meta_global": fmt(meta_s), "n_comerciais": n_com_por_super.get(sup_nome, 0),
            "prod_total": fmt(prod_s), "proj_total": fmt(proj_s),
            "pct_global": pct(proj_s, meta_s), "gap_global": gap(proj_s, meta_s),
            "meta_banco_total": fmt(meta_banco_s),
            "prod_banco_total": fmt(prod_banco_s),
            "proj_banco_total": fmt(proj_banco_s),
            "pct_banco": pct(proj_banco_s, meta_banco_s),
            "bancos": bancos_sup, "regionais": regionais_out,
        })

    # ── Empresa ───────────────────────────────────────────────────────────────
    supers_norm  = [s for s in supers_out if not s["eh_dinho"]]
    dinho_obj    = next((s for s in supers_out if s["eh_dinho"]), None)

    prod_sem  = sum(s["prod_total"] or 0 for s in supers_norm)
    prod_din  = dinho_obj["prod_total"] if dinho_obj else 0
    prod_com  = prod_sem + prod_din
    proj_sem  = prod_sem * fator
    proj_din  = prod_din * fator
    proj_com  = prod_com * fator
    meta_emp  = sum(meta_super.values())

    bancos_emp = []
    for b in bancos:
        ps  = sum(next((x["prod"] or 0 for x in s["bancos"] if x["nome"] == b["nome"]), 0)
                  for s in supers_norm)
        pd_ = next((x["prod"] or 0 for x in dinho_obj["bancos"] if x["nome"] == b["nome"]), 0) if dinho_obj else 0
        pc  = ps + pd_
        pjs = ps * fator
        pjc = pc * fator
        bancos_emp.append({
            "nome": b["nome"], "meta": fmt(b["meta"]),
            "prod_sem_dinho": fmt(ps), "proj_sem_dinho": fmt(pjs),
            "pct_sem_dinho": pct(pjs, b["meta"]), "gap_sem_dinho": gap(pjs, b["meta"]),
            "prod_com_dinho": fmt(pc), "proj_com_dinho": fmt(pjc),
            "pct_com_dinho": pct(pjc, b["meta"]), "gap_com_dinho": gap(pjc, b["meta"]),
        })

    mb_sem  = sum(x["prod_sem_dinho"] or 0 for x in bancos_emp)
    mb_com  = sum(x["prod_com_dinho"] or 0 for x in bancos_emp)
    pbj_sem = mb_sem * fator
    pbj_com = mb_com * fator

    empresa = {
        "meta_global_total":  fmt(meta_emp),
        "prod_sem_dinho":     fmt(prod_sem),
        "proj_sem_dinho":     fmt(proj_sem),
        "pct_sem_dinho":      pct(proj_sem, meta_emp),
        "gap_sem_dinho":      gap(proj_sem, meta_emp),
        "prod_com_dinho":     fmt(prod_com),
        "proj_com_dinho":     fmt(proj_com),
        "pct_com_dinho":      pct(proj_com, meta_emp),
        "gap_com_dinho":      gap(proj_com, meta_emp),
        "prod_dinho":         fmt(prod_din),
        "proj_dinho":         fmt(proj_din),
        # totais unificados (inclui Dinho) — usados pela aba Meta Global
        "prod_total":         fmt(prod_com),
        "proj_total":         fmt(proj_com),
        "pct_total":          pct(proj_com, meta_emp),
        "gap_total":          gap(proj_com, meta_emp),
        "meta_banco_total":        fmt(total_meta_banco),
        "prod_banco_sem_dinho":    fmt(mb_sem),
        "proj_banco_sem_dinho":    fmt(pbj_sem),
        "pct_banco_sem_dinho":     pct(pbj_sem, total_meta_banco),
        "prod_banco_com_dinho":    fmt(mb_com),
        "proj_banco_com_dinho":    fmt(pbj_com),
        "pct_banco_com_dinho":     pct(pbj_com, total_meta_banco),
        "bancos":             bancos_emp,
    }

    data_historico_args = (mes_str, mg_df, supers_out, fator, prod_all)

    data = {
        "info": {
            "mes_ref":          mes_str,
            "mes_label":        f"{MESES_PT[mes-1]} {ano}",
            "dias_uteis_total": du_total,
            "dias_uteis_passados": du_passados,
            "fator_proj":       round(fator, 4),
            "n_comerciais_meta":  n_com_total,
            "n_comerciais_banco": n_com_banco_total,
            "banco_mes":          banco_mes,
            "banco_mes_label":    f"{MESES_PT[banco_mes_n-1]} {banco_ano_n}",
            "banco_congelado":    banco_mes != mes_str,
            "gerado_em":        datetime.now().strftime("%d/%m/%Y %H:%M"),
        },
        "bancos_meta": [b["nome"] for b in bancos],
        "empresa":     empresa,
        "supers":      supers_out,
    }

    data["historico"] = carregar_historico(*data_historico_args)

    # ── Carteira (gestão comercial) ───────────────────────────────────────────
    _cart_params = dict(
        prod_all=prod_all, mes_str=mes_str, fator=fator,
        col_sup=col_sup, col_reg=col_reg, col_com=col_com,
        col_ban=col_ban, col_con=col_con, col_tip=col_tip,
        col_prd=col_prd, col_par=col_par, col_sta=col_sta,
    )
    data["carteira"] = _build_carteira(**_cart_params)
    # Per-super carteiras (usadas nos arquivos por superintendente)
    for sup_obj in data["supers"]:
        sup_obj["_carteira"] = _build_carteira(filter_sup=sup_obj["nome"], **_cart_params)

    # Resumo executivo por escopo (aba 📋 Resumo)
    try:
        import resumo_exec
        resumo_exec.anexar(data)
    except Exception as _e:
        print(f"  [RESUMO] AVISO: {_e}")

    return data, rows_com, rows_sum, bancos, meta_super, n_com_por_super, \
           du_total, du_passados, fator, total_meta_banco, n_com_total

# ─────────────────────────────────────────────────────────────────────────────
# Carteira — Gestão Comercial
# ─────────────────────────────────────────────────────────────────────────────

def _build_carteira(prod_all, mes_str, fator,
                    col_sup, col_reg, col_com, col_ban,
                    col_con, col_tip, col_prd, col_par, col_sta,
                    filter_sup=None):
    """Constrói o bloco 'carteira' com comparativos mensais, breakdown por
    banco/convenio/produto e hierarquia super→regional→comercial com churn."""

    def _mes_prev(ms, n=1):
        a, m = int(ms[:4]), int(ms[5:])
        for _ in range(n):
            m -= 1
            if m == 0:
                a -= 1; m = 12
        return f"{a:04d}-{m:02d}"

    def _label(ms):
        a, m = int(ms[:4]), int(ms[5:])
        return f"{MESES_PT[m-1]} {a}"

    mes_ant = _mes_prev(mes_str, 1)
    mes_pen = _mes_prev(mes_str, 2)

    # ── Filtra e prepara cada período ──────────────────────────────────────────
    def _prep(ms):
        df = prod_all[prod_all["_mes_str"] == ms].copy()
        if df.empty:
            return df
        df["_s"]  = df[col_sup].apply(clean_name) if col_sup else ""
        df["_r"]  = df[col_reg].apply(clean_name) if col_reg else ""
        df["_c"]  = df[col_com].apply(clean_name) if col_com else ""
        df["_v"]  = pd.to_numeric(df[col_prd], errors="coerce").fillna(0)
        if col_ban: df["_b"]  = df[col_ban].astype(str).str.strip()
        if col_con: df["_cn"] = df[col_con].astype(str).str.strip()
        if col_tip: df["_tp"] = df[col_tip].astype(str).str.strip()
        if col_par: df["_p"]  = df[col_par].astype(str).str.strip()
        if col_sta: df["_st"] = df[col_sta].astype(str).str.strip()
        if filter_sup:
            df = df[df["_s"] == filter_sup]
        return df

    p_pen = _prep(mes_pen)
    p_ant = _prep(mes_ant)
    p_atu = _prep(mes_str)

    def _s(df): return float(df["_v"].sum()) if not df.empty else 0.0
    def _pj(v): return round(v * fator, 2)
    def _pct(n, d): return round((n/d - 1)*100, 1) if d and d > 0 else None

    s_pen = _s(p_pen); s_ant = _s(p_ant); s_atu = _s(p_atu)
    s_proj = _pj(s_atu)

    # Dados gerais de parceiros
    def _par_set(df):
        if df.empty or "_p" not in df.columns: return set()
        return set(df[df["_v"] > 0]["_p"].dropna().unique())

    pars_pen = _par_set(p_pen)
    pars_ant = _par_set(p_ant)
    pars_atu = _par_set(p_atu)
    churn_set = pars_ant - pars_atu
    novos_set = pars_atu - pars_ant

    def _tot_ativos(df):
        if df.empty or "_st" not in df.columns or "_p" not in df.columns: return 0
        return int(df[df["_st"].str.upper().str.contains("ATIVO", na=False)]["_p"].nunique())

    resumo = {
        "mes_penultimo_label": _label(mes_pen),
        "mes_anterior_label":  _label(mes_ant),
        "mes_atual_label":     _label(mes_str),
        "prod_penultimo": fmt(s_pen),
        "prod_anterior":  fmt(s_ant),
        "prod_atual":     fmt(s_atu),
        "proj_atual":     fmt(s_proj),
        "pct_ant_pen":    _pct(s_ant, s_pen),
        "pct_proj_ant":   _pct(s_proj, s_ant),
        "n_parceiros_atu":    len(pars_atu),
        "n_parceiros_ant":    len(pars_ant),
        "n_novos":            len(novos_set),
        "n_churn":            len(churn_set),
        "n_ativos":           _tot_ativos(p_atu) or len(pars_atu),
    }

    # ── Breakdown por dimensão (banco / convenio / produto) ────────────────────
    def _por_dim(col_key):
        """col_key: '_b', '_cn', '_tp' — internal column names."""
        if not col_key:
            return []
        all_k = set()
        for df in [p_ant, p_atu]:
            if not df.empty and col_key in df.columns:
                all_k |= set(df[df["_v"] > 0][col_key].dropna().astype(str).str.strip().unique())
        all_k -= {"nan", "NAN", "", "None", "NONE"}
        result = []
        for k in all_k:
            def _fv(df, ck=col_key, kv=k):
                if df.empty or ck not in df.columns: return 0.0
                return float(df[df[ck].astype(str).str.strip() == kv]["_v"].sum())
            v_pen=_fv(p_pen); v_ant=_fv(p_ant); v_atu=_fv(p_atu); v_proj=_pj(v_atu)
            result.append({"nome":k,"pen":fmt(v_pen),"ant":fmt(v_ant),
                           "atu":fmt(v_atu),"proj":fmt(v_proj),"pct":_pct(v_proj,v_ant)})
        result.sort(key=lambda x: -(x["ant"] or 0))
        return result

    por_banco    = _por_dim("_b"  if col_ban else None)
    por_convenio = _por_dim("_cn" if col_con else None)
    por_produto  = _por_dim("_tp" if col_tip else None)

    # ── n_par por dimensão (qtd de parceiros ativos naquele banco/convenio/produto) ──
    def _add_n_par(lst, col_k):
        if not col_k or col_k not in p_atu.columns or '_p' not in p_atu.columns:
            return
        cnt = p_atu[p_atu['_v'] > 0].groupby(col_k)['_p'].nunique().to_dict()
        for item in lst:
            item['n_par'] = int(cnt.get(item['nome'], 0))

    _add_n_par(por_banco,    '_b'  if col_ban else None)
    _add_n_par(por_convenio, '_cn' if col_con else None)
    _add_n_par(por_produto,  '_tp' if col_tip else None)

    # ── Breakdown por dimensão × nível hierárquico (super / regional / comercial) ──
    def _compute_dim_maps():
        maps = {}
        for dk, col_k, have in [('b','_b',col_ban),('cn','_cn',col_con),('tp','_tp',col_tip)]:
            if not have:
                maps[dk] = {}; continue
            acc = {}
            # Inclui p_pen para mostrar todos os meses na hierarquia
            for df, period in [(p_pen,'pen'),(p_ant,'ant'),(p_atu,'atu')]:
                if df.empty or col_k not in df.columns: continue
                sub = df[df['_v'] > 0].copy()
                sub[col_k] = sub[col_k].astype(str).str.strip()
                sub = sub[~sub[col_k].isin(['nan','NAN','None','NONE',''])]
                if sub.empty: continue
                for (s, dv), v in sub.groupby(['_s', col_k])['_v'].sum().items():
                    k = ('s', str(s), str(dv))
                    acc.setdefault(k, {'pen':0.0,'ant':0.0,'atu':0.0})[period] += float(v)
                if '_r' in sub.columns:
                    for (s, r, dv), v in sub.groupby(['_s','_r',col_k])['_v'].sum().items():
                        k = ('r', str(s), str(r), str(dv))
                        acc.setdefault(k, {'pen':0.0,'ant':0.0,'atu':0.0})[period] += float(v)
                if '_r' in sub.columns and '_c' in sub.columns:
                    valid = sub[sub['_c'].astype(str).str.strip().apply(lambda x: x not in ('nan','NAN',''))]
                    if not valid.empty:
                        for (s, r, c, dv), v in valid.groupby(['_s','_r','_c',col_k])['_v'].sum().items():
                            k = ('c', str(s), str(r), str(c), str(dv))
                            acc.setdefault(k, {'pen':0.0,'ant':0.0,'atu':0.0})[period] += float(v)
            final = {}
            for k, d in acc.items():
                final[k] = {'pen': fmt(d['pen']), 'ant': fmt(d['ant']), 'atu': fmt(d['atu']), 'proj': fmt(_pj(d['atu']))}
            maps[dk] = final
        return maps

    dim_maps = _compute_dim_maps()

    # ── Cruzamento entre dimensões (banco×convenio, banco×produto, convenio×produto) ──
    def _cross_dim(col_k1, col_k2):
        """Agrega pen/ant/proj por par de dimensões para cross-filtering no JS."""
        if not col_k1 or not col_k2:
            return []
        acc = {}
        bad = {'nan','NAN','None','NONE',''}
        for df, period in [(p_pen,'pen'),(p_ant,'ant'),(p_atu,'atu')]:
            if df.empty or col_k1 not in df.columns or col_k2 not in df.columns: continue
            sub = df[df['_v'] > 0].copy()
            sub[col_k1] = sub[col_k1].astype(str).str.strip()
            sub[col_k2] = sub[col_k2].astype(str).str.strip()
            sub = sub[~sub[col_k1].isin(bad) & ~sub[col_k2].isin(bad)]
            if sub.empty: continue
            for (k1, k2), v in sub.groupby([col_k1, col_k2])['_v'].sum().items():
                key = (str(k1), str(k2))
                acc.setdefault(key, {'pen':0.0,'ant':0.0,'atu':0.0})[period] += float(v)
        return [{'k1':k1,'k2':k2,'pen':fmt(d['pen']),'ant':fmt(d['ant']),'atu':fmt(d['atu']),'proj':fmt(_pj(d['atu']))}
                for (k1,k2),d in acc.items()]

    cross_b_cv  = _cross_dim('_b'  if col_ban else None, '_cn' if col_con else None)
    cross_b_tp  = _cross_dim('_b'  if col_ban else None, '_tp' if col_tip else None)
    cross_cv_tp = _cross_dim('_cn' if col_con else None, '_tp' if col_tip else None)

    def _node_dims(level, *keys):
        """Retorna {dim_b:{val:{ant,proj}}, dim_cn:{...}, dim_tp:{...}} para um nó da hierarquia."""
        result = {}
        prefix = (level,) + tuple(str(k) for k in keys)
        n = len(prefix)
        for dk, label in [('b','dim_b'),('cn','dim_cn'),('tp','dim_tp')]:
            if not dim_maps.get(dk): continue
            nd = {k[n]: v for k, v in dim_maps[dk].items() if k[:n] == prefix}
            if nd: result[label] = nd
        return result

    # ── Hierarquia super → regional → comercial ────────────────────────────────
    from collections import defaultdict

    def _ddict3():
        return defaultdict(lambda: {"pen":0.0,"ant":0.0,"atu":0.0,
                                     "pars":defaultdict(lambda: {"pen":0.0,"ant":0.0,"atu":0.0,"st":"Ativo"})})
    def _ddict2():
        return defaultdict(lambda: {"pen":0.0,"ant":0.0,"atu":0.0,"coms":_ddict3()})
    def _ddict1():
        return defaultdict(lambda: {"pen":0.0,"ant":0.0,"atu":0.0,"regs":_ddict2()})

    hier = _ddict1()

    for df, period in [(p_pen,"pen"),(p_ant,"ant"),(p_atu,"atu")]:
        if df.empty: continue
        for row in df.to_dict('records'):
            s=str(row.get("_s","")).strip(); r=str(row.get("_r","")).strip()
            c=str(row.get("_c","")).strip(); v=float(row.get("_v",0))
            par=str(row.get("_p","")).strip() if col_par else ""
            st =str(row.get("_st","Ativo")).strip() if col_sta else "Ativo"
            if not s or s=="nan": continue
            hier[s][period]+=v; hier[s]["regs"][r][period]+=v
            if c and c!="nan":
                hier[s]["regs"][r]["coms"][c][period]+=v
                if par and par not in ("nan",""):
                    hier[s]["regs"][r]["coms"][c]["pars"][par][period]+=v
                    hier[s]["regs"][r]["coms"][c]["pars"][par]["st"]=st

    def _to_list():
        supers=[]
        for s_n,s_d in hier.items():
            s_proj=_pj(s_d["atu"])
            s_obj={"nome":s_n,"pen":fmt(s_d["pen"]),"ant":fmt(s_d["ant"]),
                   "atu":fmt(s_d["atu"]),"proj":fmt(s_proj),"pct":_pct(s_proj,s_d["ant"]),"regionais":[]}
            s_obj.update(_node_dims('s', s_n))
            for r_n,r_d in s_d["regs"].items():
                r_proj=_pj(r_d["atu"])
                r_obj={"nome":r_n,"pen":fmt(r_d["pen"]),"ant":fmt(r_d["ant"]),
                       "atu":fmt(r_d["atu"]),"proj":fmt(r_proj),"pct":_pct(r_proj,r_d["ant"]),"comerciais":[]}
                r_obj.update(_node_dims('r', s_n, r_n))
                for c_n,c_d in r_d["coms"].items():
                    c_proj=_pj(c_d["atu"])
                    pars=[{"nome":p_n,"st":p_d["st"],
                           "pen":fmt(p_d["pen"]),"ant":fmt(p_d["ant"]),
                           "atu":fmt(p_d["atu"]),"proj":fmt(_pj(p_d["atu"])),
                           "churn":(p_d["ant"]or 0)>0 and (p_d["atu"]or 0)==0}
                          for p_n,p_d in c_d["pars"].items()]
                    pars.sort(key=lambda x:-(x["atu"]or 0))
                    com_entry={
                        "nome":c_n,"pen":fmt(c_d["pen"]),"ant":fmt(c_d["ant"]),
                        "atu":fmt(c_d["atu"]),"proj":fmt(c_proj),"pct":_pct(c_proj,c_d["ant"]),
                        "n_par":sum(1 for p in pars if (p["atu"] or 0) > 0),
                        "n_churn":sum(1 for p in pars if p["churn"]),
                        "parceiros":pars}
                    com_entry.update(_node_dims('c', s_n, r_n, c_n))
                    r_obj["comerciais"].append(com_entry)
                r_obj["comerciais"].sort(key=lambda x:-(x["atu"]or 0))
                s_obj["regionais"].append(r_obj)
            s_obj["regionais"].sort(key=lambda x:-(x["atu"]or 0))
            supers.append(s_obj)
        supers.sort(key=lambda x:-(x["atu"]or 0))
        return supers

    supers_list = _to_list()

    # ── Churn detalhado ────────────────────────────────────────────────────────
    churn_list=[]
    for s in supers_list:
        for r in s["regionais"]:
            for c in r["comerciais"]:
                for p in c["parceiros"]:
                    if p["churn"]:
                        churn_list.append({"parceiro":p["nome"],"status":p["st"],
                            "comercial":c["nome"],"regional":r["nome"],"super":s["nome"],
                            "prod_pen":p["pen"],"prod_ant":p["ant"]})
    churn_list.sort(key=lambda x:-(x["prod_ant"]or 0))

    # ── Cadastro de parceiros (HIERARQUIA COMERCIAL.xlsx) ────────────────────
    cadastro_list = []
    try:
        if HIER_XLSX.exists():
            _hc = pd.read_excel(HIER_XLSX)
            _hc.columns = [str(c).strip() for c in _hc.columns]
            _cc  = next((c for c in _hc.columns if 'comercial' in c.lower()), None)
            _cr  = next((c for c in _hc.columns if 'regional' in c.lower() or 'reginal' in c.lower()), None)
            _csu = next((c for c in _hc.columns if 'superinten' in c.lower()), None)
            _cst = next((c for c in _hc.columns if 'status' in c.lower()), None)
            if _cc and _cr and _csu:
                _hc['_s'] = _hc[_csu].apply(clean_name)
                _hc['_r'] = _hc[_cr].apply(clean_name)
                _hc['_c'] = _hc[_cc].apply(clean_name)
                if _cst:
                    _hc['_atv'] = (_hc[_cst].astype(str).str.strip().str.lower() == 'ativo').astype(int)
                else:
                    _hc['_atv'] = 0
                _g = _hc.groupby(['_s','_r','_c']).agg(cad=('_atv','size'), atv=('_atv','sum')).reset_index()
                cadastro_list = [{"s":row['_s'],"r":row['_r'],"c":row['_c'],
                                  "cad":int(row['cad']),"atv":int(row['atv'])}
                                 for row in _g.to_dict('records')]
    except Exception as _e:
        print(f"  [CART] AVISO cadastro: {_e}")

    return {
        "resumo":       resumo,
        "por_banco":    por_banco,
        "por_convenio": por_convenio,
        "por_produto":  por_produto,
        "cross_b_cv":   cross_b_cv,
        "cross_b_tp":   cross_b_tp,
        "cross_cv_tp":  cross_cv_tp,
        "supers":       supers_list,
        "churn":        churn_list,
        "cadastro":     cadastro_list,
    }

# ─────────────────────────────────────────────────────────────────────────────
# Histórico anual
# ─────────────────────────────────────────────────────────────────────────────

def carregar_historico(mes_atual_str, mg_df, supers_out, fator, prod_all):
    """Percorre todos os 12 meses do ano:
       - meses passados com dados: produção real (do arquivo único prod_all)
       - mês atual: projeção calculada em supers_out
       - meses futuros: exibe só a meta (sem produção)
    """
    ano       = int(mes_atual_str[:4])
    historico = []

    # Grupos de produção por mês a partir do arquivo único
    grupos = {}
    if "_mes_str" in prod_all.columns:
        for k, v in prod_all.groupby("_mes_str"):
            grupos[k] = v

    col_prd = next((c for c in prod_all.columns if "Valor" in c or "Soma" in c), None)
    col_sup = next((c for c in prod_all.columns if "Super" in c or "super" in c), None)
    col_reg = next((c for c in prod_all.columns if "Regional" in c), None)
    col_com = next((c for c in prod_all.columns if c == "Comercial"), None)

    for mes in range(1, 13):
        mes_str  = f"{ano:04d}-{mes:02d}"
        mes_abr  = MESES_ABR[mes - 1]
        is_atual = (mes_str == mes_atual_str)
        is_futuro = (mes_str > mes_atual_str)

        # Meta do mês a partir de meta_global_2026
        meta_t, meta_por_sup = 0.0, {}
        _mg_col0 = mg_df.columns[0] if not mg_df.empty else None
        for r in mg_df.to_dict('records'):
            nome_raw = clean_name(str(r[_mg_col0]).strip())
            val = pd.to_numeric(r.get(mes_abr), errors="coerce")
            if pd.notna(val) and val > 0:
                meta_t += float(val)
                meta_por_sup[nome_raw] = float(val)

        if meta_t == 0:
            continue   # mês sem meta configurada

        # Normaliza chave do Dinho para bater com o nome da produção (ex: "DINHO SILVA")
        dinho_sup_nome = next((s["nome"] for s in supers_out if s.get("eh_dinho")), None)
        dinho_meta_key = next((k for k in list(meta_por_sup)
                               if "DINHO" in k.upper() or "SALA" in k.upper()), None)
        if dinho_sup_nome and dinho_meta_key and dinho_sup_nome != dinho_meta_key:
            meta_por_sup[dinho_sup_nome] = meta_por_sup.pop(dinho_meta_key)

        if is_atual:
            prod_t = sum(s["prod_total"] or 0 for s in supers_out)
            proj_t = prod_t * fator
            por_super = {
                s["nome"]: {
                    "meta": s["meta_global"], "prod": s["prod_total"],
                    "proj": s["proj_total"],  "pct_prod": None,
                    "pct_proj": s["pct_global"],
                }
                for s in supers_out
            }
            historico.append({
                "mes": mes_str, "mes_label": MESES_PT[mes-1][:3],
                "meta_total": fmt(meta_t), "prod_total": fmt(prod_t),
                "proj_total": fmt(proj_t), "pct_prod": None,
                "pct_proj": pct(proj_t, meta_t),
                "is_atual": True, "is_futuro": False, "por_super": por_super,
            })

        elif is_futuro:
            # Sem produção ainda — exibe só a meta para contexto
            por_super = {
                nome: {"meta": fmt(v), "prod": None, "proj": None,
                       "pct_prod": None, "pct_proj": None}
                for nome, v in meta_por_sup.items()
            }
            historico.append({
                "mes": mes_str, "mes_label": MESES_PT[mes-1][:3],
                "meta_total": fmt(meta_t), "prod_total": None,
                "proj_total": None, "pct_prod": None, "pct_proj": None,
                "is_atual": False, "is_futuro": True, "por_super": por_super,
            })

        else:
            # Mês passado — usa grupos do arquivo único
            if mes_str not in grupos or not col_prd or not col_sup:
                continue
            df = grupos[mes_str].copy()
            df["_super"]    = df[col_sup].apply(lambda x: clean_name(str(x)))
            df["_prod"]     = pd.to_numeric(df[col_prd], errors="coerce").fillna(0)
            df["_eh_dinho"] = df[col_sup].apply(eh_dinho)

            por_super = {}
            for nome_raw, meta_val in meta_por_sup.items():
                if "DINHO" in nome_raw.upper() or "SALA" in nome_raw.upper():
                    mask     = df["_eh_dinho"]
                    nome_key = str(df[mask]["_super"].iloc[0]) if mask.any() else nome_raw
                else:
                    mask     = df["_super"] == nome_raw
                    nome_key = nome_raw
                prod_val = float(df[mask]["_prod"].sum())
                por_super[nome_key] = {
                    "meta": fmt(meta_val), "prod": fmt(prod_val),
                    "proj": None, "pct_prod": pct(prod_val, meta_val), "pct_proj": None,
                }

            prod_t = float(df["_prod"].sum())
            historico.append({
                "mes": mes_str, "mes_label": MESES_PT[mes-1][:3],
                "meta_total": fmt(meta_t), "prod_total": fmt(prod_t),
                "proj_total": None, "pct_prod": pct(prod_t, meta_t),
                "pct_proj": None, "is_atual": False, "is_futuro": False,
                "por_super": por_super,
            })

    # ── Quebra por REGIONAL e por COMERCIAL (para o histórico do regional/comercial) ──
    # Produção: do arquivo único, por mês. Meta: mesmo rateio do mês atual
    # (meta do super ÷ nº de comerciais do super, × comerciais do regional).
    ncom_sup, reg_to_sup, reg_ncom, reg_dinho, com_to_sup, com_tem_meta = {}, {}, {}, {}, {}, {}
    for s in supers_out:
        sc = 0
        for r in s.get("regionais", []):
            rc = sum(1 for c in r.get("comerciais", []) if c.get("tem_meta"))
            reg_to_sup[r["nome"]] = s["nome"]; reg_ncom[r["nome"]] = rc
            reg_dinho[r["nome"]] = bool(s.get("eh_dinho")); sc += rc
            for c in r.get("comerciais", []):
                com_to_sup[c["nome"]] = s["nome"]
                com_tem_meta[c["nome"]] = bool(c.get("tem_meta")) and not bool(s.get("eh_dinho"))
        ncom_sup[s["nome"]] = sc

    prod_reg_mes, prod_com_mes = {}, {}
    if col_prd and "_mes_str" in prod_all.columns:
        tmp = prod_all.copy()
        tmp["_p"] = pd.to_numeric(tmp[col_prd], errors="coerce").fillna(0)
        if col_reg:
            tmp["_r"] = tmp[col_reg].apply(lambda x: clean_name(str(x)))
            for (ms, rn), v in tmp.groupby(["_mes_str", "_r"])["_p"].sum().items():
                prod_reg_mes[(ms, rn)] = float(v)
        if col_com:
            tmp["_c"] = tmp[col_com].apply(lambda x: clean_name(str(x)))
            for (ms, cn), v in tmp.groupby(["_mes_str", "_c"])["_p"].sum().items():
                prod_com_mes[(ms, cn)] = float(v)

    for m in historico:
        ms = m["mes"]; isfut = m.get("is_futuro"); isat = m.get("is_atual")
        mps = {sup: (info.get("meta") or 0) for sup, info in (m.get("por_super") or {}).items()}
        por_reg = {}
        for reg, sup in reg_to_sup.items():
            nps = ncom_sup.get(sup, 0)
            meta_r = 0 if reg_dinho.get(reg) else ((mps.get(sup, 0) * (reg_ncom.get(reg, 0) / nps)) if nps else 0)
            prod_r = None if isfut else prod_reg_mes.get((ms, reg), 0.0)
            proj_r = (prod_r * fator) if (isat and prod_r is not None) else None
            por_reg[reg] = {
                "meta": fmt(meta_r),
                "prod": (fmt(prod_r) if prod_r is not None else None),
                "proj": (fmt(proj_r) if proj_r is not None else None),
                "pct_prod": (pct(prod_r, meta_r) if (prod_r is not None and not isat) else None),
                "pct_proj": (pct(proj_r, meta_r) if proj_r is not None else None),
            }
        por_com = {}
        for com, sup in com_to_sup.items():
            nps = ncom_sup.get(sup, 0)
            meta_c = (mps.get(sup, 0) / nps) if (nps and com_tem_meta.get(com)) else 0
            prod_c = None if isfut else prod_com_mes.get((ms, com), 0.0)
            proj_c = (prod_c * fator) if (isat and prod_c is not None) else None
            por_com[com] = {
                "meta": fmt(meta_c),
                "prod": (fmt(prod_c) if prod_c is not None else None),
                "proj": (fmt(proj_c) if proj_c is not None else None),
                "pct_prod": (pct(prod_c, meta_c) if (prod_c is not None and not isat) else None),
                "pct_proj": (pct(proj_c, meta_c) if proj_c is not None else None),
            }
        m["por_regional"] = por_reg
        m["por_comercial"] = por_com

    return historico

# ─────────────────────────────────────────────────────────────────────────────
# Consolidado Excel
# ─────────────────────────────────────────────────────────────────────────────

def estilizar_header(ws, cor_hex="1E3A5F"):
    fill = PatternFill("solid", fgColor=cor_hex)
    font = Font(bold=True, color="FFFFFF")
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = aln
    ws.row_dimensions[1].height = 30

def auto_largura(ws, max_col=80):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, max_col)

def gerar_consolidado(data, rows_com, rows_sum):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ── Parâmetros ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Parâmetros")
    params = [
        ("Mês referência",           data["info"]["mes_ref"]),
        ("Mês label",                data["info"]["mes_label"]),
        ("Dias úteis total",         data["info"]["dias_uteis_total"]),
        ("Dias úteis passados",      data["info"]["dias_uteis_passados"]),
        ("Fator projeção",           data["info"]["fator_proj"]),
        ("Comerciais com meta",      data["info"]["n_comerciais_meta"]),
        ("Gerado em",                data["info"]["gerado_em"]),
        ("Meta Global Empresa",      data["empresa"]["meta_global_total"]),
        ("Meta Banco Total",         data["empresa"]["meta_banco_total"]),
    ]
    ws.append(["Parâmetro", "Valor"])
    estilizar_header(ws)
    for row in params:
        ws.append(list(row))
    auto_largura(ws)

    # ── Empresa ───────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Empresa")
    emp = data["empresa"]
    ws2.append(["Visão","Meta Global","Prod Real","Projeção","% Ating","GAP Global",
                 "Meta Banco","Prod Banco","Proj Banco","% Banco"])
    estilizar_header(ws2)
    ws2.append(["Sem Dinho", emp["meta_global_total"], emp["prod_sem_dinho"],
                emp["proj_sem_dinho"], emp["pct_sem_dinho"], emp["gap_sem_dinho"],
                emp["meta_banco_total"], emp["prod_banco_sem_dinho"],
                emp["proj_banco_sem_dinho"], emp["pct_banco_sem_dinho"]])
    ws2.append(["Com Dinho (bônus)", emp["meta_global_total"], emp["prod_com_dinho"],
                emp["proj_com_dinho"], emp["pct_com_dinho"], emp["gap_com_dinho"],
                emp["meta_banco_total"], emp["prod_banco_com_dinho"],
                emp["proj_banco_com_dinho"], emp["pct_banco_com_dinho"]])
    auto_largura(ws2)

    # ── Superintendentes ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Superintendentes")
    ws3.append(["Superintendente","É Dinho","Comerciais c/ Meta","Meta Global",
                 "Prod Total","Proj Total","% Global","GAP Global",
                 "Meta Banco Total","Prod Banco","Proj Banco","% Banco"])
    estilizar_header(ws3)
    for s in data["supers"]:
        ws3.append([s["nome"], "Sim" if s["eh_dinho"] else "Não",
                    s["n_comerciais"], s["meta_global"],
                    s["prod_total"], s["proj_total"], s["pct_global"], s["gap_global"],
                    s["meta_banco_total"], s["prod_banco_total"],
                    s["proj_banco_total"], s["pct_banco"]])
    auto_largura(ws3)

    # ── Regionais ─────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Regionais")
    ws4.append(["Superintendente","Regional","Estratégico","Meta Global",
                 "Prod Total","Proj Total","% Global","GAP Global",
                 "Meta Banco Total","Prod Banco","Proj Banco","% Banco"])
    estilizar_header(ws4)
    for s in data["supers"]:
        for r in s["regionais"]:
            ws4.append([s["nome"], r["nome"], "Sim" if r["eh_estrategico"] else "Não",
                        r["meta_global"], r["prod_total"], r["proj_total"],
                        r["pct_global"], r["gap_global"],
                        r["meta_banco_total"], r["prod_banco_total"],
                        r["proj_banco_total"], r["pct_banco"]])
    auto_largura(ws4)

    # ── Comerciais resumo ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Comerciais")
    if rows_sum:
        ws5.append(list(rows_sum[0].keys()))
        estilizar_header(ws5)
        for r in rows_sum:
            ws5.append(list(r.values()))
    auto_largura(ws5)

    # ── Comerciais × Banco (granular) ─────────────────────────────────────────
    ws6 = wb.create_sheet("Comerciais_x_Banco")
    if rows_com:
        ws6.append(list(rows_com[0].keys()))
        estilizar_header(ws6)
        for r in rows_com:
            ws6.append(list(r.values()))
    auto_largura(ws6)

    wb.save(OUTPUT_XLS)
    print(f"  Consolidado salvo: {OUTPUT_XLS}")

# ─────────────────────────────────────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────────────────────────────────────

def _clean_data_for_json(data):
    """Remove chaves privadas (_*) antes de serializar para JSON."""
    import copy
    d = copy.deepcopy(data)
    for s in d.get("supers", []):
        s.pop("_carteira", None)
        s.pop("_resumo", None)
    return d

def _apply_dig_replacements(tpl, dig_records, dig_estrat_json, dig_periodo, login_url="index.html"):
    """Aplica todas as substituições de template (metas + digitações)."""
    return (tpl
            .replace("__LOGIN_URL__", login_url)
            .replace("__DIG_DADOS_JSON__",     json.dumps(dig_records, ensure_ascii=False))
            .replace('"__DIG_ESTRATEGICOS_JSON__"', dig_estrat_json)
            .replace("__DIG_ESTRATEGICOS_JSON__",   dig_estrat_json)
            .replace("'__DIG_PERIODO__'",      json.dumps(dig_periodo, ensure_ascii=False))
            .replace("__DIG_PERIODO__",        dig_periodo))

def gerar_html(data, dig_records, dig_estrat_json, dig_periodo, tpl=None):
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template não encontrado: {TEMPLATE}\nCrie o arquivo dashboard_template_metas.html")
    if tpl is None:
        with open(TEMPLATE, "r", encoding="utf-8") as f:
            tpl = f.read()
    auth_path = SISTEMA_DIR / "auth.json"
    auth_json_str = "[]"
    if auth_path.exists():
        with open(auth_path, "r", encoding="utf-8") as f:
            auth_json_str = f.read().strip()
    html = (_apply_dig_replacements(tpl, dig_records, dig_estrat_json, dig_periodo)
            .replace("__DADOS_JSON__", json.dumps(_clean_data_for_json(data), ensure_ascii=False))
            .replace("__AUTH_JSON__", auth_json_str))
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Dashboard salvo: {OUTPUT_HTML}")

def montar_data_super(data, sup):
    """Monta o payload filtrado de um superintendente (mesma visão do HTML individual)."""
    sup_nome = sup["nome"]

    # ── Historico filtrado para este super ────────────────────────────────
    # Nomes das regionais/comerciais DESTE super (para recortar a quebra do histórico)
    _reg_names = {r["nome"] for r in sup.get("regionais", [])}
    _com_names = {c["nome"] for r in sup.get("regionais", []) for c in r.get("comerciais", [])}
    historico_sup = []
    for mes in data["historico"]:
        sup_entry = mes.get("por_super", {}).get(sup_nome)

        if mes["is_atual"]:
            mes_sup = {
                **mes,
                "meta_total":  sup["meta_global"],
                "prod_total":  sup["prod_total"],
                "proj_total":  sup["proj_total"],
                "pct_prod":    None,
                "pct_proj":    sup["pct_global"],
                "por_super": {sup_nome: {
                    "meta": sup["meta_global"], "prod": sup["prod_total"],
                    "proj": sup["proj_total"], "pct_prod": None,
                    "pct_proj": sup["pct_global"],
                }},
            }
        elif mes["is_futuro"]:
            meta_val = sup_entry["meta"] if sup_entry else None
            if not meta_val:
                continue
            mes_sup = {
                **mes,
                "meta_total": meta_val, "prod_total": None,
                "proj_total": None, "pct_prod": None, "pct_proj": None,
                "por_super": {sup_nome: sup_entry},
            }
        else:
            if not sup_entry:
                continue
            mes_sup = {
                **mes,
                "meta_total": sup_entry["meta"],
                "prod_total": sup_entry["prod"],
                "proj_total": None,
                "pct_prod":   sup_entry["pct_prod"],
                "pct_proj":   None,
                "por_super":  {sup_nome: sup_entry},
            }
        # Recorta a quebra por regional/comercial só para os deste super
        mes_sup["por_regional"] = {k: v for k, v in (mes.get("por_regional") or {}).items() if k in _reg_names}
        mes_sup["por_comercial"] = {k: v for k, v in (mes.get("por_comercial") or {}).items() if k in _com_names}
        historico_sup.append(mes_sup)

    # ── Empresa filtrada ──────────────────────────────────────────────────
    empresa_sup = {
        "meta_global_total":       fmt(sup["meta_global"]),
        "prod_sem_dinho":          fmt(sup["prod_total"]),
        "proj_sem_dinho":          fmt(sup["proj_total"]),
        "pct_sem_dinho":           sup["pct_global"],
        "gap_sem_dinho":           sup["gap_global"],
        "prod_com_dinho":          fmt(sup["prod_total"]),
        "proj_com_dinho":          fmt(sup["proj_total"]),
        "pct_com_dinho":           sup["pct_global"],
        "gap_com_dinho":           sup["gap_global"],
        "prod_dinho":              0,
        "proj_dinho":              0,
        "prod_total":              fmt(sup["prod_total"]),
        "proj_total":              fmt(sup["proj_total"]),
        "pct_total":               sup["pct_global"],
        "gap_total":               sup["gap_global"],
        "meta_banco_total":        fmt(sup.get("meta_banco_total")),
        "prod_banco_sem_dinho":    fmt(sup.get("prod_banco_total")),
        "proj_banco_sem_dinho":    fmt(sup.get("proj_banco_total")),
        "pct_banco_sem_dinho":     sup.get("pct_banco"),
        "prod_banco_com_dinho":    fmt(sup.get("prod_banco_total")),
        "proj_banco_com_dinho":    fmt(sup.get("proj_banco_total")),
        "pct_banco_com_dinho":     sup.get("pct_banco"),
        "bancos":                  sup.get("bancos", []),
    }

    data_sup = {
        "info": {
            **data["info"],
            "n_comerciais_meta": sup["n_comerciais"],
        },
        "bancos_meta": data["bancos_meta"],
        "empresa":     empresa_sup,
        "supers":      [sup],
        "historico":   historico_sup,
        "carteira":    sup.get("_carteira", {}),
        "resumo_exec": sup.get("_resumo", {}),
    }

    return data_sup

def gerar_html_por_super(data, dig_records, dig_estrat_json, dig_periodo, tpl=None):
    """Gera um HTML individual por superintendente em super/."""
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template não encontrado: {TEMPLATE}")
    if tpl is None:
        with open(TEMPLATE, "r", encoding="utf-8") as f:
            tpl = f.read()

    # Expõe globais para o bloco interno (closure)
    global _dig_records_global, _dig_estrat_global, _dig_periodo_global
    _dig_records_global = dig_records
    _dig_estrat_global  = dig_estrat_json
    _dig_periodo_global = dig_periodo

    for sup in data["supers"]:
        sup_nome = sup["nome"]

        data_sup = montar_data_super(data, sup)

        # ── Salva em super/ para GitHub Pages ────────────────────────────────────
        nome_arquivo = re.sub(r'[^\w\s]', '', sup_nome).strip().replace(' ', '_')

        # Busca token do usuário no auth.json (torna a URL imprevisível)
        _auth_users = []
        _auth_path = SISTEMA_DIR / 'auth.json'
        if _auth_path.exists():
            try:
                _auth_users = json.loads(_auth_path.read_text(encoding='utf-8'))
            except Exception:
                pass
        _token = next((u.get('token', '') for u in _auth_users
                       if u.get('entidade') == sup_nome and u.get('token')), '')

        # Digitações filtradas para este super
        dig_sup = [r for r in _dig_records_global if r.get('superintendente') == sup_nome]

        html_super = (_apply_dig_replacements(tpl, dig_sup, _dig_estrat_global, _dig_periodo_global, "../index.html")
                      .replace("__DADOS_JSON__", json.dumps(data_sup, ensure_ascii=False))
                      .replace("__AUTH_JSON__", "[]"))

        SUPER_DIR.mkdir(exist_ok=True)
        _fname = f"{nome_arquivo}_{_token}.html" if _token else f"{nome_arquivo}.html"
        out_super = SUPER_DIR / _fname
        with open(out_super, "w", encoding="utf-8") as f:
            f.write(html_super)

        print(f"    {sup_nome} -> super/{out_super.name}")

# ─────────────────────────────────────────────────────────────────────────────
# Login + Usuários
# ─────────────────────────────────────────────────────────────────────────────

def _hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()

def carregar_usuarios() -> list:
    """Lê usuarios.json. Cria arquivo padrão se não existir."""
    if not USUARIOS_FILE.exists():
        padrao = {"usuarios": [
            {"login": "admin", "senha": "admin123", "role": "admin", "nome": "Admin"}
        ]}
        with open(USUARIOS_FILE, "w", encoding="utf-8") as f:
            json.dump(padrao, f, ensure_ascii=False, indent=2)
        print(f"  [AVISO] usuarios.json criado com usuário padrão. Configure em: {USUARIOS_FILE}")
    with open(USUARIOS_FILE, "r", encoding="utf-8") as f:
        obj = json.load(f)
    return obj.get("usuarios", [])

def gerar_login(data, usuarios: list):
    """Gera index.html — página de login — com credenciais hasheadas."""
    if not LOGIN_TEMPLATE.exists():
        print(f"  [AVISO] login_template.html não encontrado — index.html não gerado.")
        return

    with open(LOGIN_TEMPLATE, "r", encoding="utf-8") as f:
        tpl = f.read()

    # Carrega auth.json existente para preservar tokens
    auth_path = SISTEMA_DIR / "auth.json"
    existing_auth = []
    if auth_path.exists():
        try:
            existing_auth = json.loads(auth_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    def _get_token(login_val, role_val):
        """Retorna token existente ou gera novo."""
        # Para admins, todos compartilham o mesmo token
        if role_val == "admin":
            t = next((u.get("token","") for u in existing_auth if u.get("role")=="admin" and u.get("token")), "")
            return t or secrets.token_hex(4)
        # Para demais roles, token individual
        t = next((u.get("token","") for u in existing_auth if u.get("login")==login_val and u.get("token")), "")
        return t or secrets.token_hex(4)

    # Monta config de usuários com hash + arquivo de destino + token
    users_config = []
    _admin_token = None  # token compartilhado entre admins
    for u in usuarios:
        if u.get("login") == "_instrucoes":    # ignora anotações no JSON
            continue
        role  = u.get("role", "super")
        nome  = u.get("nome", u.get("login"))
        login = u.get("login", "").strip().lower()
        senha = u.get("senha", "")
        if not login or not senha:
            continue

        token = _get_token(login, role)

        if role == "admin":
            if _admin_token is None:
                _admin_token = token
            token = _admin_token  # todos os admins compartilham token
            arquivo = f"admin/index_{token}.html"
        elif role == "super":
            entidade = u.get("entidade", "")
            slug = re.sub(r"[^\w\s]", "", entidade).strip().replace(" ", "_")
            arquivo = f"super/{slug}_{token}.html"
        elif role == "regional":
            sup_ent = u.get("super_entidade", "")
            reg_ent = u.get("entidade", "").replace(" ", "+")
            sup_slug = re.sub(r"[^\w\s]", "", sup_ent).strip().replace(" ", "_")
            # Regional usa o arquivo do seu super (com token do super)
            sup_token = next((u2.get("token","") for u2 in existing_auth
                              if u2.get("entidade") == sup_ent and u2.get("token")), token)
            arquivo = f"super/{sup_slug}_{sup_token}.html?regional={reg_ent}"
        else:
            if _admin_token is None:
                _admin_token = token
            token = _admin_token
            arquivo = f"admin/index_{token}.html"

        entry = {
            "login":   login,
            "hash":    _hash_senha(senha),
            "role":    role,
            "nome":    nome,
            "arquivo": arquivo,
            "token":   token,
        }
        if u.get("entidade"):
            entry["entidade"] = u["entidade"]
        users_config.append(entry)

    # auth.json — só gera se não existir (painel admin gerencia depois)
    if not auth_path.exists():
        with open(auth_path, "w", encoding="utf-8") as f:
            json.dump(users_config, f, ensure_ascii=False, indent=2)
        print(f"  auth.json criado: {auth_path}  ({len(users_config)} usuários)")
    else:
        print(f"  auth.json já existe — mantido (gerenciar pelo painel admin).")

    # index.html usa auth.json existente para o fallback (tokens incluídos)
    fallback_users = existing_auth if existing_auth else users_config
    ano = str(date.today().year)
    html = (tpl
            .replace("__USERS_JSON__", json.dumps(fallback_users, ensure_ascii=False))
            .replace("__ANO__", ano))
    with open(LOGIN_OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Login salvo: {LOGIN_OUTPUT}  ({len(fallback_users)} usuários)")

def gerar_html_admin(data, dig_records, dig_estrat_json, dig_periodo, tpl=None):
    """Gera admin/index_TOKEN.html — dashboard completo para administrador."""
    if not TEMPLATE.exists():
        return
    if tpl is None:
        with open(TEMPLATE, "r", encoding="utf-8") as f:
            tpl = f.read()

    # Injeta auth.json no HTML para carregar sem GitHub API
    auth_path = SISTEMA_DIR / "auth.json"
    auth_json_str = "[]"
    auth_users = []
    if auth_path.exists():
        with open(auth_path, "r", encoding="utf-8") as f:
            auth_json_str = f.read().strip()
        try:
            auth_users = json.loads(auth_json_str)
        except Exception:
            pass

    # Busca token do admin no auth.json; gera novo se não existir
    admin_token = next((u.get('token', '') for u in auth_users
                        if u.get('role') == 'admin' and u.get('token')), '')
    if not admin_token:
        admin_token = secrets.token_hex(4)
        for u in auth_users:
            if u.get('role') == 'admin':
                u['token'] = admin_token
                u['arquivo'] = f'admin/index_{admin_token}.html'
        if auth_path.exists():
            auth_path.write_text(json.dumps(auth_users, ensure_ascii=False, indent=2), encoding='utf-8')

    pasta = BASE_DIR / "admin"
    pasta.mkdir(exist_ok=True)
    nome_arquivo = f"index_{admin_token}.html"
    out = pasta / nome_arquivo
    html = (_apply_dig_replacements(tpl, dig_records, dig_estrat_json, dig_periodo, "../index.html")
            .replace("__DADOS_JSON__", json.dumps(_clean_data_for_json(data), ensure_ascii=False))
            .replace("__AUTH_JSON__", auth_json_str))
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Admin dashboard salvo: admin/{nome_arquivo}")

# ─────────────────────────────────────────────────────────────────────────────
# Git push
# ─────────────────────────────────────────────────────────────────────────────

def limpar_orfaos():
    """Remove HTMLs de super/ e admin/ que nao estao referenciados no auth.json
       (tokens antigos continuariam publicos com dados desatualizados)."""
    auth_path = SISTEMA_DIR / 'auth.json'
    if not auth_path.exists():
        return
    try:
        users = json.loads(auth_path.read_text(encoding='utf-8'))
    except Exception:
        return
    validos = {str(u.get('arquivo', '')).replace('\\', '/').split('?')[0] for u in users if u.get('arquivo')}
    removidos = 0
    for pasta in (SUPER_DIR, BASE_DIR / 'admin'):
        if not pasta.exists():
            continue
        for f in pasta.glob('*.html'):
            rel = f"{pasta.name}/{f.name}"
            if rel not in validos:
                f.unlink()
                removidos += 1
                print(f"    removido: {rel} (sem usuario no auth.json)")
    if not removidos:
        print("    nenhum arquivo orfao encontrado.")

def _publicar_supabase(data, dig_records, dig_estrat_json, dig_periodo):
    """Publica payloads (admin + um por super) na tabela dashboard_cache."""
    cfg_path = SISTEMA_DIR / "supabase_config.json"
    if not cfg_path.exists():
        return
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    if "supabase.co" not in str(cfg.get("url", "")) or "COLE-AQUI" in str(cfg.get("service_role_key", "")):
        return
    import urllib.request
    from datetime import datetime as _dt

    def upsert(escopo, payload):
        body = json.dumps([{"escopo": escopo, "payload": payload,
                            "atualizado_em": _dt.now().astimezone().isoformat()}]).encode("utf-8")
        req = urllib.request.Request(
            cfg["url"].rstrip("/") + "/rest/v1/dashboard_cache?on_conflict=escopo",
            data=body, method="POST",
            headers={"apikey": cfg["service_role_key"],
                     "Authorization": f"Bearer {cfg['service_role_key']}",
                     "Content-Type": "application/json",
                     "Prefer": "resolution=merge-duplicates,return=minimal"})
        with urllib.request.urlopen(req, timeout=300) as r:
            r.read()

    estrat = json.loads(dig_estrat_json)
    upsert("admin", {"data": _clean_data_for_json(data), "dig": dig_records,
                     "estrat": estrat, "periodo": dig_periodo})
    for sup in data["supers"]:
        dig_sup = [r for r in dig_records if r.get("superintendente") == sup["nome"]]
        upsert(sup["nome"], {"data": montar_data_super(data, sup), "dig": dig_sup,
                             "estrat": estrat, "periodo": dig_periodo})
    print(f"  [SUPA] \u2713 dashboard_cache atualizado ({1 + len(data['supers'])} escopos)")

def git_push():
    """Faz git add + commit + push. Ignora se nada mudou ou se git não configurado."""
    git = ["git", "-C", str(BASE_DIR)]
    hoje = str(date.today())

    def run(cmd, **kw):
        return subprocess.run(git + cmd, capture_output=True, text=True, **kw)

    # Verifica se é um repositório git
    check = run(["rev-parse", "--is-inside-work-tree"])
    if check.returncode != 0:
        print("  [GIT] Pasta não é um repositório git — configure com:")
        print(f"       cd \"{BASE_DIR}\"")
        print(f"       git init && git remote add origin https://github.com/diegopovoas/Painel-de-metas-comerciais.git")
        print(f"       git branch -M main && git push -u origin main")
        return

    # 1. Commita o que foi gerado AGORA — precisa vir antes da reconciliação,
    #    senão arquivos não commitados impedem o merge/pull.
    run(["add", "--", "."])
    commit = run(["commit", "-m", f"Atualização metas {hoje}"])
    if "nothing to commit" in commit.stdout or "nothing to commit" in commit.stderr:
        print("  [GIT] Nenhuma mudança — push ignorado.")
        return

    # 2. Reconcilia com o remoto (caso tenha sido publicado de outra máquina).
    #    Os arquivos são 100% gerados a partir das planilhas do Drive (iguais em
    #    qualquer máquina), então em conflito o conteúdo local prevalece (-X ours).
    #    Sem force push: o histórico remoto é preservado.
    run(["fetch", "origin", "main"])
    merge = run(["merge", "-X", "ours", "--no-edit", "origin/main"])
    if merge.returncode != 0:
        run(["merge", "--abort"])
        print("  [GIT] Aviso: não consegui reconciliar com o remoto automaticamente.")
        print(f"        {merge.stderr.strip()}")

    # 3. Publica
    push = run(["push", "origin", "main"])
    if push.returncode == 0:
        print(f"  [GIT] ✓ Push realizado → GitHub Pages atualizado!")
    else:
        print(f"  [GIT] Erro no push: {push.stderr.strip()}")
        print("        Verifique se o remote está configurado e você tem acesso.")

# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        print("\n=== GERADOR DE METAS COMERCIAL — NOVA PROMOTORA ===\n")
        print("Processando dados de metas...")
        result = processar()
        data, rows_com, rows_sum = result[0], result[1], result[2]

        print("Processando digitações...")
        dig_records, dig_estrat_json, dig_periodo = processar_digitacoes()

        # Lê template uma única vez para todas as gerações de HTML
        with open(TEMPLATE, "r", encoding="utf-8") as f:
            tpl_cache = f.read()

        print("Gerando dashboard local (diretoria)...")
        gerar_html(data, dig_records, dig_estrat_json, dig_periodo, tpl=tpl_cache)

        print("Gerando consolidado.xlsx...")
        try:
            gerar_consolidado(data, rows_com, rows_sum)
        except PermissionError:
            print("  [AVISO] consolidado.xlsx esta aberto — feche o Excel e rode novamente para atualizar o Excel.")

        print("Gerando app (index.html + app.html)...")
        try:
            import gerar_app_supabase
            gerar_app_supabase.main()
        except Exception as _e:
            print(f"  [APP] AVISO: {_e}")

        if PUSH_GIT:
            print("Publicando no GitHub Pages...")
            git_push()

        print("Publicando no Supabase...")
        try:
            _publicar_supabase(data, dig_records, dig_estrat_json, dig_periodo)
        except Exception as _e:
            print(f"  [SUPA] AVISO: publicação falhou ({_e}) — GitHub Pages não foi afetado.")

        if not os.environ.get('SKIP_BROWSER'):
            print("\n[OK] Concluido! Abrindo dashboard no navegador...\n")
            _abrir_chrome(OUTPUT_HTML.as_uri())

    except FileNotFoundError as e:
        print(f"\n[ERRO] Arquivo nao encontrado: {e}")
        print("\nVerifique se os seguintes arquivos existem:")
        print("  config/calendario.xlsx")
        print("  config/meta_global_2026.xlsx")
        print("  YYYY-MM/producao.xlsx")
        print("  YYYY-MM/meta_banco.xlsx")
        if not os.environ.get('SKIP_BROWSER'):
            try: input("\nPressione Enter para fechar...")
            except: pass
        sys.exit(1)
    except BaseException as e:
        import traceback
        print(f"\n[ERRO] {type(e).__name__}: {e}")
        traceback.print_exc()
        sys.exit(1)
        try: input("\nPressione Enter para fechar...")
        except: pass
